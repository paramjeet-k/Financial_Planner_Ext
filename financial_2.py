import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Page configuration
st.set_page_config(
    page_title="Financial Planner",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header { font-size: 2.5rem; font-weight: bold; color: white; text-align: center; margin-bottom: 2rem; }
    .metric-container { background-color: #121212; padding: 1rem; border-radius: 10px; margin: 0.5rem 0; color: white; }
    .stMetric { background-color: #1e1e1e; color: white; padding: 1rem; border-radius: 8px; box-shadow: 0 2px 4px rgba(255,255,255,0.1); }
</style>
""", unsafe_allow_html=True)

# Format helpers
def format_inr(amount):
    if amount >= 1e7:
        return f"₹{amount/1e7:.2f} Cr"
    elif amount >= 1e5:
        return f"₹{amount/1e5:.2f} L"
    else:
        return f"₹{amount:,.0f}"

# Financial calculations
def calculate_compound_interest(principal, rate, time, compound_freq=12):
    return principal * (1 + rate / compound_freq) ** (compound_freq * time)

def calculate_sip_returns(monthly_investment, annual_rate, years):
    monthly_rate = annual_rate / 12
    months = years * 12
    return monthly_investment * (((1 + monthly_rate) ** months - 1) / monthly_rate)

def calculate_loan_emi(principal, rate, tenure):
    monthly_rate = rate / 12
    months = tenure * 12
    return principal * (monthly_rate * (1 + monthly_rate) ** months) / ((1 + monthly_rate) ** months - 1)

# Excel export function
def create_downloadable_excel(data_dict):
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, data in data_dict.items():
        ws = wb.create_sheet(title=sheet_name)
        if isinstance(data, pd.DataFrame):
            for col_idx, col_name in enumerate(data.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            for row_idx, row in enumerate(data.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        else:
            row = 1
            for key, value in data.items():
                ws.cell(row=row, column=1, value=key).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
                row += 1
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# Main application
def main():
    st.markdown('<h1 class="main-header">💰 Comprehensive Financial Planner</h1>', unsafe_allow_html=True)

    st.sidebar.title("🗭 Navigation")
    planning_type = st.sidebar.selectbox("Select Planning Type", [
        "Dashboard", "Investment Planning", "Retirement Planning",
        "Education Planning", "Loan Planning", "Budget Planning",
        "Emergency Fund Planning"
    ])

    if 'financial_data' not in st.session_state:
        st.session_state.financial_data = {}

    if planning_type == "Dashboard":
        dashboard()
    elif planning_type == "Investment Planning":
        investment()
    elif planning_type == "Retirement Planning":
        retirement()
    elif planning_type == "Education Planning":
        education()
    elif planning_type == "Loan Planning":
        loan()
    elif planning_type == "Budget Planning":
        budget()
    elif planning_type == "Emergency Fund Planning":
        emergency()

    # Download report
    st.sidebar.markdown("---")
    st.sidebar.subheader("📅 Download Reports")
    if st.sidebar.button("Download Excel Report"):
        if st.session_state.financial_data:
            excel_data = create_downloadable_excel(st.session_state.financial_data)
            st.sidebar.download_button(
                label="📊 Download Excel File",
                data=excel_data,
                file_name=f"financial_plan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.sidebar.warning("No data available to download. Run some calculations first.")

# Dashboard Page
def dashboard():
    st.header("📊 Financial Dashboard")
    col1, col2, col3, col4 = st.columns(4)
    with col1: st.metric("Total Assets", "₹0", "0%")
    with col2: st.metric("Total Liabilities", "₹0", "0%")
    with col3: st.metric("Net Worth", "₹0", "0%")
    with col4: st.metric("Monthly Savings", "₹0", "0%")

    if st.button("Generate Sample Portfolio"):
        data = {
            'Asset Class': ['Equity MF', 'Debt MF', 'PPF', 'FD', 'Gold', 'Real Estate'],
            'Value': [350000, 250000, 150000, 100000, 50000, 100000]
        }
        df = pd.DataFrame(data)
        df['Formatted'] = df['Value'].apply(format_inr)
        fig = px.pie(df, values='Value', names='Asset Class', hover_data=['Formatted'], title='Portfolio Allocation')
        st.plotly_chart(fig, use_container_width=True)

# Investment Planner
def investment():
    st.header("📈 Investment Planning")
    col1, col2 = st.columns(2)
    with col1:
        investment_type = st.selectbox("Investment Type", ["Lump Sum", "SIP", "Both"])
        lump = st.number_input("Lump Sum (₹)", 0, value=500000) if investment_type != "SIP" else 0
        sip = st.number_input("Monthly SIP (₹)", 0, value=10000) if investment_type != "Lump Sum" else 0
        rate = st.slider("Expected Return (%)", 1.0, 30.0, 12.0)
        years = st.slider("Years", 1, 50, 10)
        if st.button("Calculate Investment Returns"):
            lump_future = calculate_compound_interest(lump, rate/100, years)
            sip_future = calculate_sip_returns(sip, rate/100, years)
            invested = lump + (sip * 12 * years)
            final_value = lump_future + sip_future
            gains = final_value - invested
            st.session_state.financial_data['investment'] = {
                'Total Invested': format_inr(invested),
                'Future Value': format_inr(final_value),
                'Total Gains': format_inr(gains),
                'CAGR': f"{((final_value/invested-1)*100):.2f}%"
            }
    with col2:
        if 'investment' in st.session_state.financial_data:
            for k, v in st.session_state.financial_data['investment'].items():
                st.metric(k, v)
            # Plot
            years_range = list(range(years + 1))
            df = pd.DataFrame({
                'Year': years_range,
                'Lump Sum': [calculate_compound_interest(lump, rate/100, y) for y in years_range],
                'SIP': [calculate_sip_returns(sip, rate/100, y) for y in years_range]
            })
            df['Total'] = df['Lump Sum'] + df['SIP']
            fig = px.line(df, x='Year', y=['Lump Sum', 'SIP', 'Total'], title="Growth Over Time")
            st.plotly_chart(fig, use_container_width=True)

# Add other planners here (retirement, education, loan, etc.)

def retirement():
    st.header("🏖️ Retirement Planning")
    st.info("Feature coming soon!")

def education():
    st.header("🎓 Education Planning")
    st.info("Feature coming soon!")

def loan():
    st.header("🏠 Loan Planning")
    st.info("Feature coming soon!")

def budget():
    st.header("📊 Budget Planning")
    st.info("Feature coming soon!")

def emergency():
    st.header("🚨 Emergency Fund Planning")
    st.info("Feature coming soon!")

# Run app
if __name__ == "__main__":
    main()
